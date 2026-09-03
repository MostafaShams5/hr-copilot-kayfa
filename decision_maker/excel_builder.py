import io
import pandas as pd
from typing import List, Optional
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from handoff import DecisionOutput, DecisionRecommendation, InterviewOutput

class ExcelReportBuilder:
    @staticmethod
    def generate_dossier(decisions: List[DecisionOutput], candidates: Optional[List[InterviewOutput]] = None) -> io.BytesIO:
        cand_map = {c.candidate_id: c for c in candidates} if candidates else {}
        
        rows = []
        for d in decisions:
            cand = cand_map.get(d.candidate_id)
            rec_val = d.recommendation.value if hasattr(d.recommendation, "value") else str(d.recommendation)
            
            row = {
                "Candidate ID": d.candidate_id,
                "Recommendation": rec_val,
                "Final Vector Score (%)": round(d.final_score, 1),
                "Base Composite Score (%)": round(d.base_score, 1),
                "Technical Score (%)": round(cand.technical_score, 1) if cand else "N/A",
                "Behavioral / HR Score (%)": round(cand.hr_score, 1) if cand else "N/A",
                "Gate Status": cand.gate_status if cand else "N/A",
                "Hard Veto Reason": d.veto_reason or "None",
                "Confidence Penalty": "Yes" if d.ai_confidence_penalty_applied else "No",
                "Evaluation Confidence": f"{round(cand.evaluation_confidence * 100, 1)}%" if cand else "N/A",
                "Validated Strengths": " • " + " • ".join(cand.strengths) if (cand and cand.strengths) else "None reported",
                "Remaining Gaps": " • " + " • ".join(cand.remaining_gaps) if (cand and cand.remaining_gaps) else "None reported",
                "Behavioral Red Flags": " • " + " • ".join(cand.behavioral_red_flags) if (cand and cand.behavioral_red_flags) else "None detected",
                "AI Executive Reasoning": cand.reasoning if cand else "N/A",
                "Hiring Manager Probing Questions": d.probing_questions if d.probing_questions else "None",
                "Decided At (UTC)": str(d.decided_at)[:19] if d.decided_at else "N/A"
            }
            rows.append(row)

        df = pd.DataFrame(rows)
        if not df.empty and "Final Vector Score (%)" in df.columns:
            df = df.sort_values(by="Final Vector Score (%)", ascending=False)

        approved = df[df["Recommendation"] == DecisionRecommendation.APPROVE.value]
        escalated = df[df["Recommendation"] == DecisionRecommendation.ESCALATE.value]
        rejected = df[df["Recommendation"] == DecisionRecommendation.REJECT.value]

        output = io.BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Full Master Dossier", index=False)
            approved.to_excel(writer, sheet_name="Shortlist (Hire)", index=False)
            escalated.to_excel(writer, sheet_name="Manual Review", index=False)
            rejected.to_excel(writer, sheet_name="Rejected", index=False)
            
            # Styling headers and auto-adjusting column widths
            header_fill = PatternFill(start_color="4652D3", end_color="4652D3", fill_type="solid")
            header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            thin_border = Border(
                left=Side(style='thin', color='E2E8F0'),
                right=Side(style='thin', color='E2E8F0'),
                top=Side(style='thin', color='E2E8F0'),
                bottom=Side(style='thin', color='E2E8F0')
            )

            for sheet_name in writer.sheets:
                ws = writer.sheets[sheet_name]
                ws.views.sheetView[0].showGridLines = True
                
                # Format header row
                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

                # Format data cells and adjust column widths
                for col in ws.columns:
                    col_letter = col[0].column_letter
                    max_len = max((len(str(cell.value or '')) for cell in col), default=10)
                    ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)
                    for cell in col[1:]:
                        cell.border = thin_border
                        cell.alignment = Alignment(vertical="center")

        output.seek(0)
        return output
